import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference

def main():
    # Create the Inputs sheet
    inputs_data = {
        "Item": [
            "Units Sold (Lettuce)",
            "Price per Crate",
            "Labor Cost per Unit",
            "Fertilizer Cost",
            "Equipment Rent",
            "Overhead"
        ],
        "Unit": ["crates", "USD", "USD", "USD/month", "USD/month", "USD/month"],
        "Monthly Forecast Value": [10000, 2.5, 0.4, 3000, 5000, 2000]
    }
    inputs_df = pd.DataFrame(inputs_data)

    # Create the Budget sheet
    months = [
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
    ]
    units_sold = [10000] * 12
    price = 2.5
    labor_cost_per_unit = 0.4
    fertilizer = 3000
    rent = 5000
    overhead = 2000

    budget_rows = []
    for month in months:
        revenue = price * 10000
        labor_cost = 10000 * labor_cost_per_unit
        total_cost = labor_cost + fertilizer + rent + overhead
        gross_profit = revenue - total_cost
        budget_rows.append([month, 10000, revenue, labor_cost, fertilizer, rent, overhead, total_cost, gross_profit])

    budget_df = pd.DataFrame(budget_rows, columns=[
        "Month", "Units Sold", "Revenue", "Labor Cost", "Fertilizer",
        "Rent", "Overhead", "Total Cost", "Gross Profit"
    ])

    # Create the Actuals & Variance sheet
    import numpy as np

    actual_rows = []
    for i in range(12):
        actual_revenue = budget_df.loc[i, "Revenue"] * np.random.uniform(0.85, 1.15) # type: ignore
        actual_cost = budget_df.loc[i, "Total Cost"] * np.random.uniform(0.85, 1.15) # type: ignore
        budget_revenue = budget_df.loc[i, "Revenue"]
        budget_cost = budget_df.loc[i, "Total Cost"]
        budget_gross_profit = budget_df.loc[i, "Gross Profit"]
        revenue_var = actual_revenue - budget_revenue # type: ignore
        revenue_var_pct = revenue_var / budget_revenue if budget_revenue != 0 else 0 # type: ignore
        cost_var = actual_cost - budget_cost # type: ignore
        margin_diff = (budget_gross_profit / budget_revenue if budget_revenue != 0 else 0) - (((actual_revenue - actual_cost) / actual_revenue) if actual_revenue != 0 else 0) # type: ignore
        actual_rows.append([
            months[i], budget_revenue, actual_revenue, revenue_var,
            revenue_var_pct, budget_cost, actual_cost, cost_var, margin_diff
        ])

    actuals_df = pd.DataFrame(actual_rows, columns=[
        "Month", "Budget Revenue", "Actual Revenue", "Variance ($)", "Variance (%)",
        "Budget Cost", "Actual Cost", "Cost Variance ($)", "Margin Diff (%)"
    ])

    # Write to Excel
    wb = Workbook()
    ws_inputs = wb.active
    ws_inputs.title = "Inputs" # type: ignore

    for r in dataframe_to_rows(inputs_df, index=False, header=True):
        ws_inputs.append(r) # type: ignore

    ws_budget = wb.create_sheet(title="Budget")
    for r in dataframe_to_rows(budget_df, index=False, header=True):
        ws_budget.append(r)

    ws_actuals = wb.create_sheet(title="Actuals & Variance")
    for r in dataframe_to_rows(actuals_df, index=False, header=True):
        ws_actuals.append(r)
    
    # cosmetics
    for sheet in [ws_inputs, ws_budget, ws_actuals]:
        # Apply bold font to headers
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        max_length = 0
        # Adjust column widths
        for column in sheet.columns:
            if column[0] is None:
                continue
            column_letter = column[0].column_letter
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # create chart
    ws = wb["Actuals & Variance"]

    chart = LineChart()
    chart.title = "Monthly Actuals vs Budget"
    chart.style = 13
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Revenue ($)"

    cats = Reference(ws, min_col=1, min_row=2, max_row=13)  # Months

    data = Reference(ws, min_col=2, min_row=1, max_row=13, max_col=3)  # Actuals and Budget
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "A16")  # Place chart at A16 # type: ignore

    # Save to file
    file_path = "../data/Budget_vs_Actuals_Template.xlsx"
    wb.save(file_path)

    print(f"Template created successfully at {file_path}")

    return 

if __name__ == "__main__":
    main()